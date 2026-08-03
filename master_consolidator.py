"""
LPH Master Database Consolidation Engine
=========================================
Reads every worksheet of every xlsx/xlsm/xls/csv file in a source folder,
normalizes inconsistent headers onto one shared schema tailored to LPH's
real-estate owner-contact data, auto-extends that schema when genuinely new
fields appear, and appends everything into one growing Master Database --
with zero data loss, full logging, and non-destructive data enrichment.

Design principles:
  - Every worksheet is read. Only truly empty sheets are skipped.
  - Header normalization is synonym-based and CONFIGURABLE (HEADER_SYNONYMS
    below). Unknown headers are never dropped -- they become new columns.
  - Headerless sheets (first row is already data) are detected and given
    inferred column names instead of losing that row as a fake header.
  - One bad sheet/file never stops the run: errors are caught per-sheet,
    logged, and processing continues.
  - Master Database is append-only across runs: existing rows/columns are
    never deleted, reordered, or overwritten; new columns are appended.
  - Original values are NEVER modified. Enrichment (phone normalization,
    email validation, cross-file duplicate detection) is added as new,
    separate columns -- the raw value stays exactly as supplied.
  - A processed-file registry (content hash + sheet name) prevents
    re-ingesting the same sheet twice across repeated runs.
  - A post-write size/row-count guard flags abnormal master-file growth in
    the run summary so a future runaway-growth bug gets caught in a
    Telegram notification instead of surfacing later as an unopenable
    100MB+ file in Excel.
  - Runs headless from n8n's Execute Command node: takes a JSON config,
    prints exactly one JSON summary line to stdout, and uses exit codes
    (0 = clean, 1 = completed with warnings/errors logged, 2 = fatal) so
    the workflow can branch on success/failure without parsing prose.

Extending later (AI header mapping, fuzzy matching, address normalization,
incremental cloud sync, etc.) is a matter of adding functions in the
PLUGIN POINTS section without touching the core loop.
"""

import argparse
import csv
import hashlib
import json
import re
import sys
import time
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

# ============================================================================
# CONFIG -- header synonym map, tailored to LPH real-estate owner-contact
# data. Add new synonyms here; no code changes needed elsewhere. Keys are
# canonical column names as they appear in the Master Database.
# ============================================================================
HEADER_SYNONYMS = {
    "Name": ["name", "owner name", "owner", "client name", "owners name",
             "customer name", "name of owner", "client", "full name",
             "first name", "client full name", "primary applicant name",
             "nameen", "joint acct name", "account name"],
    "Community": ["community", "community name", "master location"],
    "Sub-Community": ["sub community", "subcommunity", "sub-community"],
    "Building/Cluster": ["building", "tower", "building name", "building 1",
                          "buildingname 2", "buildingnameen", "tower name",
                          "bldg.", "bldg. no.", "cluster", "building/cluster",
                          "building cluster"],
    "Unit Number": ["unit", "unit no", "unit number", "villa number", "villa no",
                     "property number", "property no", "no of unit", "no of units",
                     "unit id", "unitnumber", "flat number", "flat", "unit name"],
    "Size": ["size", "actual size", "unit size", "area", "actual area"],
    "Plot Reg. No": ["plot reg no", "plot reg. no", "plot registration number",
                       "reg no", "registration number", "regis"],
    "Plot Number": ["plot number", "plot no", "land number", "land no", "landnumber", "plotno"],
    "DMNO": ["dm no", "dmno", "municipality number", "municipality no"],
    "DMsubno": ["dm sub no", "dmsubno", "municipality sub no", "municipality subno",
                 "dm sub number"],
    "Bedroom": ["bhk", "no bhk", "bedrooms", "bedroom", "no of bedrooms", "bed", "beds",
                 "rooms", "rooms description"],
    "Type (Buyer/Seller)": ["type buyer seller", "buyer seller", "buyer/seller",
                              "type (buyer/seller)", "transaction type", "party type"],
    "Email Address": ["email", "e-mail", "email address", "e mail", "email add"],
    "PI number": ["pi number", "pi no", "pino", "pi num"],
    "Nationality": ["nationality", "nation"],
    "Property Type": ["property type", "unit type", "sub type", "flat typology"],
    "Date": ["date", "transaction date", "procedure date", "date of transaction"],
    "Procedure Value": ["procedure value", "procedurevalue", "transaction amount",
                          "transaction value", "amount"],
    "Developer": ["developer", "project developer"],
    "Project": ["project", "project name", "master project", "emaar project",
                "master project land", "project lnd", "sub project"],
    "Serial No": ["serial no", "serial number", "sno", "sr no"],
    "Emirates ID Number": ["idnumber", "uaeidnumber", "emirates id number"],
    "Passport Number": ["passport"],
    "Date of Birth": ["birthdate", "dob"],
    "Gender": ["gender"],
}

MAIN_HEADERS = [
    "Name", "Community", "Sub-Community", "Building/Cluster", "Unit Number",
    "Size", "Plot Reg. No", "Plot Number", "DMNO", "DMsubno", "Bedroom",
    "Type (Buyer/Seller)", "Mobile 1", "Mobile 2", "Mobile 3",
    "Email Address", "PI number", "Nationality", "Property Type", "Date",
    "Procedure Value", "Developer", "Project",
]

PHONE_SYNONYMS_RAW = [
    "phone", "mobile", "number", "contact number", "phone number",
    "contact no", "contact", "mobile no", "mobile number", "tel",
    "telephone", "phone no", "primary phone", "phone mobile", "mobile 1",
    "primary mobile number", "poa mobile no.", "secondary mobile",
    "secondary phone", "alternate number", "alt number",
    "alternative number", "second contact", "other number", "mobile 2",
    "poa phone no.", "mobile no.3", "mobile phone3", "mobile 3", "phone 2",
    "phone no.3", "telephone number", "telephone residence",
    "telephone office", "phone 1", "general",
]

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv"}

META_COLUMNS = ["Record ID", "_source_file", "_source_sheet", "_source_row", "_ingested_at"]
ENRICHMENT_COLUMNS = ["Mobile Number (Normalized)", "Mobile Country", "Mobile Number Valid",
                       "Email Valid", "Possible Duplicate Of"]

MAX_SAFE_MASTER_SIZE_MB = 50
MAX_SAFE_ROW_COUNT = 200_000


# ============================================================================
# HEADER NORMALIZATION
# ============================================================================
def normalize_header(raw):
    if raw is None:
        return ""
    s = str(raw)
    s = unicodedata.normalize("NFKD", s)
    s = s.replace("\n", " ").replace("\t", " ").replace("\r", " ")
    s = s.lower()
    s = re.sub(r"[`'’‘]", "", s)
    s = re.sub(r"[,.]", "", s)
    s = re.sub(r"[\-_/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


PHONE_SYNONYMS_NORMALIZED = {normalize_header(s) for s in PHONE_SYNONYMS_RAW}


class SchemaRegistry:
    def __init__(self, existing_columns=None):
        self.columns = list(existing_columns) if existing_columns else []
        self._norm_to_canonical = {}
        for canon, synonyms in HEADER_SYNONYMS.items():
            for syn in synonyms:
                self._norm_to_canonical[normalize_header(syn)] = canon
            self._norm_to_canonical[normalize_header(canon)] = canon
        for col in self.columns:
            self._norm_to_canonical.setdefault(normalize_header(col), col)
        for main_header in MAIN_HEADERS:
            self.ensure_column(main_header)

    def canonical_for(self, raw_header):
        nk = normalize_header(raw_header)
        if not nk:
            return None
        if nk in self._norm_to_canonical:
            canon = self._norm_to_canonical[nk]
        else:
            canon = re.sub(r"\s+", " ", str(raw_header).strip())
            self._norm_to_canonical[nk] = canon
        self.ensure_column(canon)
        return canon

    def ensure_column(self, canon):
        if canon not in self.columns:
            self.columns.append(canon)


_KNOWN_HEADER_NORMS = set()
for _canon, _syns in HEADER_SYNONYMS.items():
    _KNOWN_HEADER_NORMS.add(normalize_header(_canon))
    for _s in _syns:
        _KNOWN_HEADER_NORMS.add(normalize_header(_s))


_EMAIL_LOOSE_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")

SKIP_SHEET_NAME_SUBSTRINGS = ["instruction", "readme", "notes", "summary", "dashboard"]

_PIVOT_MARKER_NORMS = {"row labels", "column labels", "grand total"}


def _sheet_name_excluded(sheet_name):
    n = str(sheet_name or "").strip().lower()
    return any(sub in n for sub in SKIP_SHEET_NAME_SUBSTRINGS)


def _looks_like_pivot_table(rows, max_scan=5):
    for row in rows[:max_scan]:
        for c in row:
            if c is None:
                continue
            norm = normalize_header(c)
            if norm in _PIVOT_MARKER_NORMS or norm.startswith("count of"):
                return True
    return False


def _looks_like_data_value(s):
    s = s.strip()
    if _EMAIL_LOOSE_RE.search(s):
        return True
    digits_only = re.sub(r"[^\d]", "", s)
    if digits_only and len(digits_only) >= 4 and len(digits_only) >= len(s) - 3:
        return True
    if re.match(r"^[A-Za-z0-9]+(-[A-Za-z0-9]+){2,}$", s):
        return True
    words = s.split()
    if len(words) >= 2 and all(w.isalpha() for w in words):
        return True
    return False


def looks_like_header(row):
    non_empty = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
    if len(non_empty) < 2:
        return False
    if any(_EMAIL_LOOSE_RE.search(c) for c in non_empty):
        return False
    if any(re.sub(r"[\s]", "", c).isdigit() for c in non_empty):
        return False
    if any(normalize_header(c) in _KNOWN_HEADER_NORMS for c in non_empty):
        return True
    label_like = sum(1 for c in non_empty if not _looks_like_data_value(c))
    return label_like >= max(1, len(non_empty) * 0.6)


def find_header_row(rows, max_scan=10):
    for i, row in enumerate(rows[:max_scan]):
        if looks_like_header(row):
            return i
    return None


_BEDROOM_LIKE_RE = re.compile(r"\b(bedroom|studio|\bbr\b|bhk)\b", re.IGNORECASE)

KNOWN_SUB_COMMUNITY_NAMES = {"richmond", "topanga"}

KNOWN_VIEW_FACING_VALUES = {"front", "back", "side", "corner",
                              "sea view", "garden view", "pool view",
                              "park view", "community view", "boulevard view"}


def infer_generic_headers(sample_rows, width):
    headers = [None] * width
    for col in range(width):
        vals = [r[col] for r in sample_rows[:30]
                 if col < len(r) and r[col] is not None and str(r[col]).strip() != ""]
        if not vals:
            headers[col] = f"Column {col + 1}"
            continue
        n = len(vals)
        digit_like = sum(1 for v in vals
                          if str(v).replace(" ", "").replace("+", "").isdigit()
                          and len(str(v).replace(" ", "")) >= 7)
        email_like = sum(1 for v in vals if _EMAIL_LOOSE_RE.search(str(v)))
        code_like = sum(1 for v in vals
                          if re.match(r"^[A-Za-z0-9/\-]{4,}$", str(v).strip().replace(" ", ""))
                          and any(ch.isdigit() for ch in str(v)))
        bedroom_like = sum(1 for v in vals if isinstance(v, str) and _BEDROOM_LIKE_RE.search(v))
        community_like = sum(1 for v in vals
                               if isinstance(v, str) and v.strip().lower() in KNOWN_SUB_COMMUNITY_NAMES)
        view_like = sum(1 for v in vals
                          if isinstance(v, str) and v.strip().lower() in KNOWN_VIEW_FACING_VALUES)
        alpha_multiword = sum(1 for v in vals
                               if isinstance(v, str) and len(v.split()) >= 2
                               and v.replace(" ", "").isalpha())
        if email_like / n > 0.6:
            headers[col] = "Email Address"
        elif digit_like / n > 0.6:
            headers[col] = "Mobile 1"
        elif bedroom_like / n > 0.6:
            headers[col] = "Bedrooms"
        elif community_like / n > 0.6:
            headers[col] = "Community"
        elif view_like / n > 0.6:
            headers[col] = "View / Facing"
        elif alpha_multiword / n > 0.6:
            headers[col] = "Name"
        elif code_like / n > 0.6:
            headers[col] = "Unit Number"
        else:
            headers[col] = f"Column {col + 1}"
    return headers


# ============================================================================
# NON-DESTRUCTIVE ENRICHMENT (PLUGIN POINTS)
# ============================================================================
import phonenumbers
from phonenumbers import NumberParseException

COUNTRY_NAMES = {
    "AE": "UAE", "SA": "Saudi Arabia", "IQ": "Iraq", "IR": "Iran",
    "IN": "India", "PK": "Pakistan", "EG": "Egypt", "JO": "Jordan",
    "LB": "Lebanon", "SY": "Syria", "KW": "Kuwait", "QA": "Qatar",
    "BH": "Bahrain", "OM": "Oman", "YE": "Yemen", "GB": "UK",
    "US": "USA", "CA": "Canada", "AU": "Australia", "FR": "France",
    "DE": "Germany", "PH": "Philippines", "BD": "Bangladesh",
    "NG": "Nigeria", "CN": "China", "RU": "Russia", "TR": "Turkey",
    "ZA": "South Africa", "KE": "Kenya", "MA": "Morocco", "TN": "Tunisia",
    "DZ": "Algeria", "LK": "Sri Lanka", "NP": "Nepal", "AF": "Afghanistan",
    "SD": "Sudan", "ET": "Ethiopia", "MY": "Malaysia", "ID": "Indonesia",
    "IT": "Italy", "ES": "Spain", "NL": "Netherlands", "CH": "Switzerland",
}

_GARBAGE_PHONE_VALUES = {"null", "n/a", "na", "none", "-", "0", "00", "nil"}


def normalize_phone(raw, default_region="AE"):
    if raw is None:
        return None, None, None
    s = str(raw).strip()
    if not s or s.lower() in _GARBAGE_PHONE_VALUES:
        return None, None, None
    s = s.replace("|", "")
    if s.startswith("00"):
        s = "+" + s[2:]
    try:
        parsed = phonenumbers.parse(s, None if s.startswith("+") else default_region)
    except NumberParseException:
        return None, None, None
    valid = phonenumbers.is_valid_number(parsed)
    possible = phonenumbers.is_possible_number(parsed)
    if not (valid or possible):
        return None, None, None
    e164 = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
    region = phonenumbers.region_code_for_number(parsed)
    country_name = COUNTRY_NAMES.get(region, region)
    return e164, country_name, valid


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def is_valid_email(raw):
    if raw is None or str(raw).strip() == "":
        return None
    return bool(EMAIL_RE.match(str(raw).strip()))


def enrich_record(rec):
    for slot in ("Mobile 1", "Mobile 2", "Mobile 3"):
        raw = rec.get(slot)
        if raw is None:
            continue
        e164, country_name, valid = normalize_phone(raw)
        if e164:
            rec["Mobile Number (Normalized)"] = e164
            rec["Mobile Country"] = country_name
            rec["Mobile Number Valid"] = valid
            break
    email = rec.get("Email Address")
    if email is not None:
        rec["Email Valid"] = is_valid_email(email)


def detect_duplicates(records):
    seen_phone = {}
    seen_email = {}
    for idx, rec in enumerate(records):
        matches = set()
        norm_phone = rec.get("Mobile Number (Normalized)")
        if norm_phone:
            if norm_phone in seen_phone:
                matches.add(seen_phone[norm_phone])
            else:
                seen_phone[norm_phone] = idx
        email = rec.get("Email Address")
        if email:
            key = str(email).strip().lower()
            if key in seen_email:
                matches.add(seen_email[key])
            else:
                seen_email[key] = idx
        if matches:
            rec["Possible Duplicate Of"] = ", ".join(str(m + 2) for m in sorted(matches))


# ============================================================================
# FILE READING (per-sheet, error-isolated)
# ============================================================================
def sha256_of_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def _split_header_and_data(non_empty_rows):
    if not non_empty_rows:
        return None
    if _looks_like_pivot_table(non_empty_rows):
        return "PIVOT"
    hdr_idx = find_header_row(non_empty_rows)
    if hdr_idx is None:
        width = max(len(r) for r in non_empty_rows)
        header = infer_generic_headers(non_empty_rows, width)
        data = non_empty_rows
    else:
        header = non_empty_rows[hdr_idx]
        data = non_empty_rows[hdr_idx + 1:]
    return header, data


def read_sheet(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)
        if len(rows) > 200000:
            break
    non_empty_rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    return _split_header_and_data(non_empty_rows)


def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        rows = list(csv.reader(f))
    non_empty_rows = [r for r in rows if any(c.strip() for c in r)]
    return _split_header_and_data(non_empty_rows)


def read_workbook(path, registry, log_entries, processed_registry):
    """Returns (records, file_hash). file_hash is None only when the file
    couldn't even be read (hashing failed) -- callers use this to populate
    the Master_Index / Duplicate_Report file-level metadata."""
    records = []
    ext = path.suffix.lower()
    try:
        file_hash = sha256_of_file(path)
    except Exception as e:
        log_entries.append({"file": path.name, "sheet": None, "status": "ERROR",
                             "error": f"Could not read file: {e}", "rows_imported": 0,
                             "timestamp": now_iso()})
        return records, None

    if ext == ".csv":
        sheets = [("Sheet1", None)]
    else:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            log_entries.append({"file": path.name, "sheet": None, "status": "ERROR",
                                 "error": f"Failed to open workbook: {e}", "rows_imported": 0,
                                 "timestamp": now_iso()})
            return records
        sheets = [(name, wb[name]) for name in wb.sheetnames]

    for sheet_name, ws in sheets:
        t0 = time.time()
        if _sheet_name_excluded(sheet_name):
            log_entries.append({"file": path.name, "sheet": sheet_name,
                                 "status": "SKIPPED_EXCLUDED_SHEET_NAME", "rows_imported": 0,
                                 "timestamp": now_iso(), "duration_sec": round(time.time() - t0, 3)})
            continue
        sheet_key = f"{file_hash}::{sheet_name}"
        if sheet_key in processed_registry:
            log_entries.append({"file": path.name, "sheet": sheet_name,
                                 "status": "SKIPPED_DUPLICATE", "rows_imported": 0,
                                 "timestamp": now_iso(), "duration_sec": round(time.time() - t0, 3)})
            continue
        try:
            result = read_csv(path) if ext == ".csv" else read_sheet(ws)
            if result == "PIVOT":
                log_entries.append({"file": path.name, "sheet": sheet_name,
                                     "status": "SKIPPED_PIVOT_TABLE", "rows_imported": 0,
                                     "timestamp": now_iso(), "duration_sec": round(time.time() - t0, 3)})
                continue
            if result is None:
                log_entries.append({"file": path.name, "sheet": sheet_name,
                                     "status": "EMPTY_SKIPPED", "rows_imported": 0,
                                     "timestamp": now_iso(), "duration_sec": round(time.time() - t0, 3)})
                continue
            header, data = result

            new_cols_before = set(registry.columns)
            col_map = []
            for h in header:
                if normalize_header(h) in PHONE_SYNONYMS_NORMALIZED:
                    col_map.append("__PHONE__")
                else:
                    col_map.append(registry.canonical_for(h))

            PHONE_LIKE_CANONICALS = {"Telephone Number", "Alternate Number(s)"}

            imported = 0
            for i, row in enumerate(data, start=1):
                rec = {}
                phone_vals = []
                for h_idx, canon in enumerate(col_map):
                    val = row[h_idx] if h_idx < len(row) else None
                    if val is None or str(val).strip() == "":
                        continue
                    val_str = str(val).strip()
                    if canon == "__PHONE__":
                        if val_str not in phone_vals:
                            phone_vals.append(val_str)
                        continue
                    if canon is None:
                        continue
                    if canon in rec and str(rec[canon]).strip() != val_str:
                        if canon in PHONE_LIKE_CANONICALS:
                            existing_parts = [p.strip() for p in str(rec[canon]).split(" / ")]
                            if val_str not in existing_parts:
                                rec[canon] = str(rec[canon]) + " / " + val_str
                        else:
                            suffix = 2
                            alt = f"{canon} ({suffix})"
                            while alt in rec:
                                suffix += 1
                                alt = f"{canon} ({suffix})"
                            rec[alt] = val
                            registry.ensure_column(alt)
                    else:
                        rec[canon] = val
                if phone_vals:
                    if len(phone_vals) >= 1:
                        rec["Mobile 1"] = phone_vals[0]
                    if len(phone_vals) >= 2:
                        rec["Mobile 2"] = phone_vals[1]
                    if len(phone_vals) >= 3:
                        rec["Mobile 3"] = " / ".join(phone_vals[2:])
                if not rec:
                    continue
                enrich_record(rec)
                rec["_source_file"] = path.name
                rec["_source_sheet"] = sheet_name
                rec["_source_row"] = i
                rec["_ingested_at"] = now_iso()
                records.append(rec)
                imported += 1

            new_cols = set(registry.columns) - new_cols_before
            processed_registry.add(sheet_key)
            log_entries.append({"file": path.name, "sheet": sheet_name, "status": "OK",
                                 "rows_imported": imported,
                                 "columns_mapped": len([c for c in col_map if c]),
                                 "new_columns_created": sorted(new_cols),
                                 "timestamp": now_iso(), "duration_sec": round(time.time() - t0, 3)})
        except Exception as e:
            log_entries.append({"file": path.name, "sheet": sheet_name, "status": "ERROR",
                                 "error": str(e), "rows_imported": 0, "timestamp": now_iso(),
                                 "duration_sec": round(time.time() - t0, 3)})
            continue

    return records, file_hash


# ============================================================================
# MASTER DATABASE READ/WRITE
# ============================================================================
def _is_corrupted_column_name(name):
    s = str(name).strip()
    if _EMAIL_LOOSE_RE.search(s):
        return True
    if re.sub(r"[\s]", "", s).isdigit():
        return True
    return False


def _sanitize_sheet_name(name, used_names):
    s = re.sub(r'[:\\/?*\[\]]', '-', str(name)).strip()
    if not s:
        s = "Unassigned"
    s = s[:31]
    base = s
    n = 2
    while s.lower() in used_names:
        suffix = f" ({n})"
        s = base[: 31 - len(suffix)] + suffix
        n += 1
    used_names.add(s.lower())
    return s


def _group_key_for(rec):
    for field in ("Community", "Project"):
        val = rec.get(field)
        if val is not None and str(val).strip() != "":
            return str(val).strip()
    src = rec.get("_source_file")
    if src:
        return re.sub(r"\.(xlsx|xlsm|xls|csv)$", "", str(src), flags=re.IGNORECASE).replace("_", " ").strip()
    return "Unassigned"


def write_batch_workbook(path, columns, records, group_sheets=True):
    """Writes ONE batch's consolidated data only. Output is restricted to
    EXACTLY the 23 requested columns, in this fixed order -- no Record ID,
    no _source_file/_source_sheet/_source_row/_ingested_at tracking
    columns, no enrichment columns (Mobile Number (Normalized) etc.), and
    no auto-discovered extra columns, even if the source sheets had them.
    (Those fields still exist internally on each record and are still used
    for duplicate detection -- they're just not written to this sheet.)
    """
    ordered_cols = list(MAIN_HEADERS)
    wb = openpyxl.Workbook(write_only=True)

    ws = wb.create_sheet("Consolidated")
    ws.append(ordered_cols)
    for rec in records:
        ws.append([rec.get(c, None) for c in ordered_cols])

    if group_sheets:
        groups = {}
        for rec in records:
            key = _group_key_for(rec)
            groups.setdefault(key, []).append(rec)
        used_names = {"consolidated"}
        for key in sorted(groups.keys(), key=lambda k: (-len(groups[k]), k.lower())):
            sheet_name = _sanitize_sheet_name(key, used_names)
            gws = wb.create_sheet(sheet_name)
            gws.append(ordered_cols)
            for rec in groups[key]:
                gws.append([rec.get(c, None) for c in ordered_cols])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_log(path, log_entries):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(log_entries, f, indent=2, default=str)


# ============================================================================
# BATCH ORCHESTRATION -- state, Master_Index, Duplicate_Report,
# Processing_Log.xlsx, Workflow_Summary.xlsx, crash-safe resume.
#
# Everything above this section is the ORIGINAL, UNCHANGED processing
# pipeline (header normalization, enrichment, dedup-by-content, per-sheet
# read logic). This section only wraps it with batching per the master
# prompt -- it does not alter how any individual file/sheet is parsed.
# ============================================================================
STATE_VERSION = 1


def default_state():
    return {
        "version": STATE_VERSION,
        "next_batch_number": 1,
        "next_record_id": 1,
        "workflow_start": None,
        "batches": [],          # list of per-batch summary dicts (for Workflow_Summary)
        "file_registry": {},    # file_hash -> {name, size, first_seen_batch, status}
    }


def load_state(state_path):
    if not state_path.exists():
        return default_state()
    try:
        state = json.loads(state_path.read_text())
    except Exception:
        return default_state()
    base = default_state()
    base.update(state)
    return base


def save_state(state_path, state):
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(state, indent=2, default=str))


def _load_manifest(manifest_path):
    """Optional JSON: {"filename.xlsx": {"drive_url": "...", "drive_id": "...",
    "original_path": "..."}}. n8n writes this alongside the downloaded batch
    files so Master_Index can carry a clickable Drive link + original folder
    path even though the Python engine only sees local temp files."""
    if not manifest_path or not Path(manifest_path).exists():
        return {}
    try:
        return json.loads(Path(manifest_path).read_text())
    except Exception:
        return {}


def _load_xlsx_rows(path):
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    return [dict(zip(header, r)) for r in rows[1:]
            if any(v is not None and str(v).strip() != "" for v in r)]


def _write_xlsx_rows(path, header, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(header)
    for r in rows:
        ws.append([r.get(h) for h in header])
    wb.save(path)


MASTER_INDEX_COLUMNS = [
    "Record ID", "Name", "DirectoryName", "Extension", "Size (KB)",
    "LastWriteTime", "Batch Number", "Consolidated Batch Workbook Name",
    "Processing Status", "Duplicate Status", "Duplicate Of",
    "Processing Timestamp", "Drive URL",
]

DUPLICATE_REPORT_COLUMNS = [
    "Record ID", "Name", "DirectoryName", "File Hash", "Size (KB)",
    "Duplicate Status", "Duplicate Of", "Batch Number", "Detected At",
]

PROCESSING_LOG_COLUMNS = [
    "Batch Number", "Event", "Files Processed", "Files Skipped",
    "Duplicate Files", "Errors", "Warnings", "Processing Time (sec)",
    "Timestamp",
]


def run_batch(source_dir, batch_workbook_path, state_path, master_index_path,
              dup_report_path, processing_log_path, sheet_registry_path,
              manifest_path=None):
    """Processes every supported file currently sitting in source_dir as
    ONE batch: writes a self-contained Consolidated_Batch_NNN.xlsx, appends
    this batch's rows to the running Master_Index and Duplicate_Report, and
    persists state so a crash/restart resumes at the next batch number
    instead of overwriting an already-completed one."""
    source_dir = Path(source_dir)
    batch_workbook_path = Path(batch_workbook_path)
    state_path = Path(state_path)
    master_index_path = Path(master_index_path)
    dup_report_path = Path(dup_report_path)
    processing_log_path = Path(processing_log_path)
    sheet_registry_path = Path(sheet_registry_path)

    t_batch_start = time.time()
    state = load_state(state_path)
    if state["workflow_start"] is None:
        state["workflow_start"] = now_iso()

    batch_number = state["next_batch_number"]
    manifest = _load_manifest(manifest_path)

    processed_sheet_hashes = set()
    if sheet_registry_path.exists():
        try:
            processed_sheet_hashes = set(json.loads(sheet_registry_path.read_text()))
        except Exception:
            processed_sheet_hashes = set()

    existing_index_rows = _load_xlsx_rows(master_index_path)
    existing_dup_rows = _load_xlsx_rows(dup_report_path)
    existing_log_rows = _load_xlsx_rows(processing_log_path)

    registry = SchemaRegistry()
    log_entries = []
    all_records = []
    new_index_rows = []
    new_dup_rows = []
    errors, warnings = 0, 0

    if not source_dir.exists():
        raise FileNotFoundError(f"Source directory does not exist: {source_dir}")

    files = sorted(p for p in source_dir.rglob("*")
                    if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS)

    for path in files:
        try:
            stat = path.stat()
            size_kb = round(stat.st_size / 1024, 2)
            last_write = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
        except OSError:
            size_kb, last_write = None, None

        try:
            recs, file_hash = read_workbook(path, registry, log_entries, processed_sheet_hashes)
            status = "Processed" if any(l["file"] == path.name and l["status"] == "OK"
                                          for l in log_entries) else "Processed (no new rows)"
        except Exception as e:
            recs, file_hash = [], None
            status = "Failed"
            errors += 1
            log_entries.append({"file": path.name, "sheet": None, "status": "ERROR",
                                 "error": str(e), "rows_imported": 0, "timestamp": now_iso()})

        for rec in recs:
            rec["Record ID"] = state["next_record_id"]
            state["next_record_id"] += 1
        all_records.extend(recs)

        # File-level duplicate detection (Step 7): hash first, then
        # size+filename as a fallback signal -- separate from the existing
        # person-level (phone/email) row dedup, which still runs below.
        dup_status, dup_of = "Unique", None
        if file_hash:
            prior = state["file_registry"].get(file_hash)
            if prior:
                dup_status, dup_of = "Duplicate", prior.get("record_id")
            else:
                for h, meta in state["file_registry"].items():
                    if meta.get("size_kb") == size_kb and meta.get("name") == path.name and h != file_hash:
                        dup_status, dup_of = "Possible Duplicate", meta.get("record_id")
                        break

        record_id = recs[0]["Record ID"] if recs else state["next_record_id"]
        if not recs:
            state["next_record_id"] += 1

        meta = manifest.get(path.name, {})
        index_row = {
            "Record ID": record_id,
            "Name": path.name,
            "DirectoryName": meta.get("original_path", str(path.parent)),
            "Extension": path.suffix.lower(),
            "Size (KB)": size_kb,
            "LastWriteTime": last_write,
            "Batch Number": batch_number,
            "Consolidated Batch Workbook Name": batch_workbook_path.name,
            "Processing Status": status,
            "Duplicate Status": dup_status,
            "Duplicate Of": dup_of,
            "Processing Timestamp": now_iso(),
            "Drive URL": meta.get("drive_url", ""),
        }
        new_index_rows.append(index_row)

        if file_hash:
            state["file_registry"][file_hash] = {
                "name": path.name, "size_kb": size_kb, "record_id": record_id,
                "batch_number": batch_number,
            }

        if dup_status != "Unique":
            new_dup_rows.append({
                "Record ID": record_id, "Name": path.name,
                "DirectoryName": index_row["DirectoryName"], "File Hash": file_hash,
                "Size (KB)": size_kb, "Duplicate Status": dup_status,
                "Duplicate Of": dup_of, "Batch Number": batch_number,
                "Detected At": now_iso(),
            })

    detect_duplicates(all_records)
    used_columns = [c for c in registry.columns
                     if any(r.get(c) not in (None, "") for r in all_records)]
    write_batch_workbook(batch_workbook_path, used_columns, all_records)

    _write_xlsx_rows(master_index_path, MASTER_INDEX_COLUMNS, existing_index_rows + new_index_rows)
    _write_xlsx_rows(dup_report_path, DUPLICATE_REPORT_COLUMNS, existing_dup_rows + new_dup_rows)

    sheets_errored = sum(1 for l in log_entries if l["status"] == "ERROR")
    warnings = sum(1 for l in log_entries if l["status"] in
                    ("SKIPPED_PIVOT_TABLE", "SKIPPED_EXCLUDED_SHEET_NAME"))
    duration = round(time.time() - t_batch_start, 3)

    batch_log_row = {
        "Batch Number": batch_number, "Event": "Batch Complete",
        "Files Processed": len(files), "Files Skipped": 0,
        "Duplicate Files": len(new_dup_rows), "Errors": sheets_errored,
        "Warnings": warnings, "Processing Time (sec)": duration,
        "Timestamp": now_iso(),
    }
    _write_xlsx_rows(processing_log_path, PROCESSING_LOG_COLUMNS,
                      existing_log_rows + [batch_log_row])

    sheet_registry_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_registry_path.write_text(json.dumps(sorted(processed_sheet_hashes), indent=2))

    state["batches"].append({
        "batch_number": batch_number, "files_in_batch": len(files),
        "rows_imported": len(all_records), "errors": sheets_errored,
        "duplicates": len(new_dup_rows), "duration_sec": duration,
        "completed_at": now_iso(),
    })
    state["next_batch_number"] = batch_number + 1
    save_state(state_path, state)

    summary = {
        "status": "warning" if sheets_errored > 0 else "ok",
        "batch_number": batch_number,
        "files_in_batch": len(files),
        "rows_imported": len(all_records),
        "duplicates_flagged": len(new_dup_rows),
        "sheets_errored": sheets_errored,
        "sheets_warnings": warnings,
        "batch_workbook": str(batch_workbook_path),
        "duration_sec": duration,
        "errors": [l for l in log_entries if l["status"] == "ERROR"],
    }
    return summary


def run_finalize(state_path, workflow_summary_path):
    """Closes out a run: writes Workflow_Summary.xlsx from the persisted
    state (no source files touched -- purely an aggregation step)."""
    state_path = Path(state_path)
    workflow_summary_path = Path(workflow_summary_path)
    state = load_state(state_path)
    batches = state["batches"]

    total_files = sum(b["files_in_batch"] for b in batches)
    total_rows = sum(b["rows_imported"] for b in batches)
    total_errors = sum(b["errors"] for b in batches)
    total_dupes = sum(b["duplicates"] for b in batches)
    total_time = sum(b["duration_sec"] for b in batches)
    avg_batch_time = round(total_time / len(batches), 3) if batches else 0

    summary_rows = [{
        "Total Source Files Discovered": total_files,
        "Total Batches Created": len(batches),
        "Total Processed Successfully": total_files - total_errors,
        "Total Failed": total_errors,
        "Total Duplicates": total_dupes,
        "Total Rows Imported": total_rows,
        "Total Execution Time (sec)": round(total_time, 3),
        "Average Batch Processing Time (sec)": avg_batch_time,
        "Workflow Start": state["workflow_start"],
        "Workflow Completion Timestamp": now_iso(),
    }]
    _write_xlsx_rows(workflow_summary_path, list(summary_rows[0].keys()), summary_rows)

    return {
        "status": "ok", "total_batches": len(batches), "total_files": total_files,
        "total_rows_imported": total_rows, "total_errors": total_errors,
        "total_duplicates": total_dupes, "workflow_summary_path": str(workflow_summary_path),
    }


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================
def load_config(config_path):
    cfg = json.loads(Path(config_path).read_text())
    required = ["source_dir", "batch_workbook_path", "state_path", "master_index_path",
                "dup_report_path", "processing_log_path", "sheet_registry_path"]
    missing = [k for k in required if k not in cfg]
    if missing:
        raise ValueError(f"Config missing required keys: {missing}")
    return cfg


def main():
    parser = argparse.ArgumentParser(description="LPH Master Database Consolidation Engine")
    parser.add_argument("--mode", choices=["batch", "finalize"], default="batch",
                         help="'batch' processes one batch of files sitting in --source-dir; "
                              "'finalize' writes Workflow_Summary.xlsx from persisted state "
                              "and touches no source files.")
    parser.add_argument("--config", help="Path to JSON config file (batch mode only)")
    parser.add_argument("--source-dir", help="Folder containing this batch's downloaded files")
    parser.add_argument("--batch-workbook-path", help="Output path for Consolidated_Batch_NNN.xlsx "
                                                        "('NNN' is auto-substituted with the "
                                                        "zero-padded batch number from state)")
    parser.add_argument("--state-path", required=False, help="Path to batch_state.json (persisted "
                                                               "batch counter, record IDs, file "
                                                               "registry -- crash-safe resume)")
    parser.add_argument("--master-index-path", help="Path to Master_Index.xlsx")
    parser.add_argument("--dup-report-path", help="Path to Duplicate_Report.xlsx")
    parser.add_argument("--processing-log-path", help="Path to Processing_Log.xlsx")
    parser.add_argument("--sheet-registry-path", help="Path to processed_sheet_hashes.json "
                                                        "(existing content-level dedup, unchanged)")
    parser.add_argument("--manifest-path", help="Optional JSON mapping filename -> "
                                                 "{drive_url, original_path} for Master_Index")
    parser.add_argument("--workflow-summary-path", help="Path to Workflow_Summary.xlsx "
                                                          "(finalize mode only)")
    args = parser.parse_args()

    try:
        if args.mode == "finalize":
            if not args.state_path or not args.workflow_summary_path:
                print(json.dumps({"status": "fatal_error",
                                   "error": "finalize mode requires --state-path and --workflow-summary-path"}))
                sys.exit(2)
            summary = run_finalize(args.state_path, args.workflow_summary_path)
            print(json.dumps(summary))
            sys.exit(0)

        if args.config:
            cfg = load_config(args.config)
        else:
            cfg = {
                "source_dir": args.source_dir,
                "batch_workbook_path": args.batch_workbook_path,
                "state_path": args.state_path,
                "master_index_path": args.master_index_path,
                "dup_report_path": args.dup_report_path,
                "processing_log_path": args.processing_log_path,
                "sheet_registry_path": args.sheet_registry_path,
                "manifest_path": args.manifest_path,
            }

        missing = [k for k in ("source_dir", "batch_workbook_path", "state_path",
                                "master_index_path", "dup_report_path",
                                "processing_log_path", "sheet_registry_path")
                   if not cfg.get(k)]
        if missing:
            print(json.dumps({"status": "fatal_error",
                               "error": f"batch mode missing required args: {missing}"}))
            sys.exit(2)

        batch_workbook_path = cfg["batch_workbook_path"]
        if "NNN" in batch_workbook_path:
            state = load_state(Path(cfg["state_path"]))
            batch_workbook_path = batch_workbook_path.replace("NNN", f"{state['next_batch_number']:03d}")

        summary = run_batch(
            source_dir=cfg["source_dir"],
            batch_workbook_path=batch_workbook_path,
            state_path=cfg["state_path"],
            master_index_path=cfg["master_index_path"],
            dup_report_path=cfg["dup_report_path"],
            processing_log_path=cfg["processing_log_path"],
            sheet_registry_path=cfg["sheet_registry_path"],
            manifest_path=cfg.get("manifest_path"),
        )
        print(json.dumps(summary))
        sys.exit(1 if summary["sheets_errored"] > 0 else 0)

    except Exception as e:
        print(json.dumps({"status": "fatal_error", "error": str(e)}))
        sys.exit(2)


if __name__ == "__main__":
    main()
